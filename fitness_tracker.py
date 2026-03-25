import openpyxl
from openpyxl import Workbook

# Create a new workbook
wb = Workbook()

# --- Workout Tracker Sheet ---
ws1 = wb.active
ws1.title = "Workout Tracker"
ws1.append(["Date", "Day (Push/Pull/Leg)", "Exercise", "Sets", "Reps", "Weight (kg)", "Remarks"])

# --- Measurements Sheet ---
ws2 = wb.create_sheet(title="Measurements")
ws2.append(["Date", "Body Weight (kg)", "Body Fat %", "Neck (in)", "Chest (in)", "Waist (in)",
            "Hips (in)", "Arms (in)", "Forearms (in)", "Thighs (in)", "Calves (in)", "Remarks"])

# --- Diet Tracker Sheet ---
ws3 = wb.create_sheet(title="Diet Tracker")
ws3.append(["Date", "Meal (Breakfast/Lunch/Dinner/Snack)", "Food Item", "Quantity", "Calories (kcal)",
            "Protein (g)", "Carbs (g)", "Fats (g)", "Fiber (g)", "Sugar (g)", "Sodium (mg)",
            "Vitamins", "Minerals", "Water Intake (ml)", "Remarks"])

# Save the file
wb.save("Fitness_Tracker.xlsx")
print("✅ Fitness_Tracker.xlsx created successfully!")

